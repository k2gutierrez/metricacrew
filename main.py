import streamlit as st
import pandas as pd
import numpy as np
import os
import sys
import tempfile
from datetime import date
from typing import Dict, List, Tuple
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

st.set_page_config(page_title="Valuador DCF - Simulador Profesional", page_icon="🏢", layout="wide", initial_sidebar_state="expanded")

AZUL_OSCURO = "FF1B2A4A"
AZUL_MEDIO = "FF2E86C1"
AZUL_CLARO = "FFD6EAF8"
AMARILLO_INPUT = "FFFFF9C4"
VERDE_OUTPUT = "FFD5F5E3"
NARANJA_ACENTO = "FFF39C12"
COLOR_FONDO = "FFF8F9FA"
COLOR_TEXTO = "FF2C3E50"

ANOS = ["Año -1", "Año 0", "Año 1", "Año 2", "Año 3", "Año 4", "Año 5"]
ANOS_PROY = ["Año 1", "Año 2", "Año 3", "Año 4", "Año 5"]

SAMPLE_DATA = {
    "empresa": "Valor, S.A. de C.V.",
    "fecha": date(2025, 1, 1),
    "moneda": "MXN",
    "ingresos": {"Año -1": 622590, "Año 0": 741120},
    "crecimiento": {"Año -1": 0.163, "Año 0": 0.190, "Año 1": 0.050, "Año 2": 0.050, "Año 3": 0.050, "Año 4": 0.050, "Año 5": 0.050},
    "costo_directo_pct": {"Año -1": 0.636, "Año 0": 0.636, "Año 1": 0.636, "Año 2": 0.631, "Año 3": 0.631, "Año 4": 0.626, "Año 5": 0.626},
    "gastos_venta_pct": {"Año -1": 0.235, "Año 0": 0.242, "Año 1": 0.242, "Año 2": 0.242, "Año 3": 0.242, "Año 4": 0.242, "Año 5": 0.242},
    "gastos_admin_pct": {"Año -1": 0.019, "Año 0": 0.018, "Año 1": 0.019, "Año 2": 0.019, "Año 3": 0.019, "Año 4": 0.019, "Año 5": 0.019},
    "dya_pct": {"Año -1": 0.026, "Año 0": 0.025, "Año 1": 0.025, "Año 2": 0.025, "Año 3": 0.025, "Año 4": 0.025, "Año 5": 0.025},
    "dias_cxc": 45, "dias_inventario": 20, "dias_cxp": 69,
    "isr_ptu": 0.008, "capex_pct": 0.025,
    "wacc": 0.175, "crecimiento_terminal": 0.0, "metodo_tv": "Growth",
    "deuda": 54739, "efectivo": 7411, "capital_contable": 218956,
}

# Función auxiliar para convertir a float de forma segura
def safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


class DCFEngine:
    """Motor de cálculo para valuación por Flujos de Caja Descontados (DCF)."""

    def __init__(self):
        pass

    def calcular_valuacion(self, params: dict) -> dict:
        """Calcula la valuación completa usando DCF a 5 años con valor terminal."""
        df_er = self._proyectar_estado_resultados(params)
        df_wc, delta_wc = self._calcular_capital_trabajo(params, df_er)
        fcfs, df_fcf = self._calcular_fcf(params, df_er, delta_wc)
        terminal_value = self._calcular_valor_terminal(
            fcfs[-1], params["wacc"],
            params.get("crecimiento_terminal", 0.0),
            params.get("metodo_tv", "Growth")
        )
        fcf_descontados = self._descontar_flujos(fcfs, params["wacc"])
        num_anios_proy = len(ANOS_PROY)
        tv_descontado = terminal_value / ((1 + params["wacc"]) ** num_anios_proy)
        ev = sum(fcf_descontados) + tv_descontado
        equity = ev - params.get("deuda", 0) + params.get("efectivo", 0)
        multiples = self._calcular_multiples(ev, df_er, equity, params)
        return {
            "estado_resultados": df_er, "capital_trabajo": df_wc, "flujo_efectivo": df_fcf,
            "fcfs": fcfs, "fcf_descontados": fcf_descontados,
            "terminal_value": terminal_value, "tv_descontado": tv_descontado,
            "enterprise_value": ev, "equity_value": equity,
            "multiples": multiples, "wacc": params["wacc"],
            "crecimiento_terminal": params.get("crecimiento_terminal", 0.0),
            "metodo_tv": params.get("metodo_tv", "Growth"),
        }

    def _proyectar_estado_resultados(self, params: dict) -> pd.DataFrame:
        """Proyecta el Estado de Resultados para los años -1, 0, 1, 2, 3, 4, 5."""
        ingresos_base = params["ingresos_base"]
        crecimiento = params["crecimiento"]
        todos_anios = ANOS
        data = {}
        ingresos = {}
        if "Año -1" in params.get("ingresos_historicos", {}):
            ingresos["Año -1"] = params["ingresos_historicos"]["Año -1"]
        else:
            ingresos["Año -1"] = ingresos_base / (1 + crecimiento.get("Año 0", 0.19))
        ingresos["Año 0"] = ingresos_base
        for i in range(1, 6):
            anio = f"Año {i}"
            anio_ant = f"Año {i-1}"
            tasa = crecimiento.get(anio, 0.05)
            ingresos[anio] = ingresos[anio_ant] * (1 + tasa)
        data["Ingresos"] = ingresos
        crec_dict = {}
        for i, anio in enumerate(todos_anios):
            if i == 0:
                crec_dict[anio] = None
            else:
                ant = todos_anios[i-1]
                crec_dict[anio] = (ingresos[anio] / ingresos[ant]) - 1
        data["Crecimiento %"] = crec_dict
        costo_directo = {}
        cp = params.get("costo_directo_pct", {})
        for a in todos_anios:
            pct = cp.get(a, 0.636)
            costo_directo[a] = ingresos[a] * pct
        data["Costo Directo"] = costo_directo
        gastos_venta = {}
        gvp = params.get("gastos_venta_pct", {})
        for a in todos_anios:
            pct = gvp.get(a, 0.242)
            gastos_venta[a] = ingresos[a] * pct
        data["Gastos de Venta"] = gastos_venta
        gastos_admin = {}
        gap = params.get("gastos_admin_pct", {})
        for a in todos_anios:
            pct = gap.get(a, 0.019)
            gastos_admin[a] = ingresos[a] * pct
        data["Gastos Administrativos"] = gastos_admin
        dya = {}
        dyap = params.get("dya_pct", {})
        for a in todos_anios:
            pct = dyap.get(a, 0.025)
            dya[a] = ingresos[a] * pct
        data["D&A"] = dya
        costos_totales = {}
        for a in todos_anios:
            costos_totales[a] = costo_directo.get(a, 0) + gastos_venta.get(a, 0) + gastos_admin.get(a, 0) + dya.get(a, 0)
        data["Total Costos y Gastos"] = costos_totales
        ebit = {}
        for a in todos_anios:
            ebit[a] = ingresos[a] - costos_totales[a]
        data["EBIT"] = ebit
        isr_ptu = {}
        tasa_isr = params.get("isr_ptu", 0.008)
        for a in todos_anios:
            isr_ptu[a] = ebit[a] * tasa_isr
        data["ISR + PTU"] = isr_ptu
        nopat = {}
        for a in todos_anios:
            nopat[a] = ebit[a] - isr_ptu[a]
        data["NOPAT"] = nopat
        utilidad_neta = {}
        for a in todos_anios:
            utilidad_neta[a] = nopat[a]
        data["Utilidad Neta"] = utilidad_neta
        return pd.DataFrame(data, index=todos_anios).T

    def _calcular_capital_trabajo(self, params: dict, df_er: pd.DataFrame) -> Tuple[pd.DataFrame, List[float]]:
        """Calcula el Capital de Trabajo y su variación anual."""
        anios = df_er.columns.tolist()
        dias_cxc = params.get("dias_cxc", 45)
        dias_inv = params.get("dias_inventario", 20)
        dias_cxp = params.get("dias_cxp", 69)
        data = {}
        cxc = {}
        for a in anios:
            ing = df_er.loc["Ingresos", a]
            cxc[a] = (dias_cxc / 365) * ing
        data["Cuentas por Cobrar"] = cxc
        inventario = {}
        for a in anios:
            cd = df_er.loc["Costo Directo", a]
            inventario[a] = (dias_inv / 365) * cd
        data["Inventario"] = inventario
        cxp = {}
        for a in anios:
            ct = df_er.loc["Costo Directo", a] + df_er.loc["Gastos de Venta", a] + df_er.loc["Gastos Administrativos", a]
            cxp[a] = (dias_cxp / 365) * ct
        data["Cuentas por Pagar"] = cxp
        wc = {}
        for a in anios:
            wc[a] = cxc[a] + inventario[a] - cxp[a]
        data["Working Capital (WC)"] = wc
        delta_wc_list = []
        for i, a in enumerate(anios):
            if i == 0:
                delta_wc_list.append(0.0)
            else:
                delta_wc_list.append(wc[a] - wc[anios[i-1]])
        delta_wc = {}
        for i, a in enumerate(anios):
            delta_wc[a] = delta_wc_list[i]
        data["Δ Working Capital"] = delta_wc
        return pd.DataFrame(data, index=anios).T, delta_wc_list

    def _calcular_fcf(self, params: dict, df_er: pd.DataFrame, delta_wc: List[float]) -> Tuple[List[float], pd.DataFrame]:
        """Calcula los Flujos de Caja Libre (FCF)."""
        anios = df_er.columns.tolist()
        capex_pct = params.get("capex_pct", 0.025)
        data = {}
        nopat = {}
        for a in anios:
            nopat[a] = df_er.loc["NOPAT", a]
        data["NOPAT"] = nopat
        dya = {}
        for a in anios:
            dya[a] = df_er.loc["D&A", a]
        data["(+) D&A"] = dya
        d_wc = {}
        for i, a in enumerate(anios):
            d_wc[a] = delta_wc[i]
        data["(-) Δ Capital Trabajo"] = d_wc
        capex = {}
        for a in anios:
            capex[a] = df_er.loc["Ingresos", a] * capex_pct
        data["(-) CapEx"] = capex
        fcf_list = []
        for i, a in enumerate(anios):
            fcf_val = nopat[a] + dya[a] - d_wc[a] - capex[a]
            fcf_list.append(fcf_val)
        fcf = {}
        for i, a in enumerate(anios):
            fcf[a] = fcf_list[i]
        data["Flujo de Caja Libre (FCF)"] = fcf
        return fcf_list, pd.DataFrame(data, index=anios).T

    def _calcular_valor_terminal(self, ultimo_fcf: float, wacc: float, g_terminal: float = 0.0, metodo: str = "Growth") -> float:
        """Calcula el Valor Terminal (Perpetuidad)."""
        if metodo == "None":
            return 0.0
        elif metodo == "Zero":
            return ultimo_fcf / wacc if wacc > 0 else 0.0
        elif metodo == "Growth":
            if wacc <= g_terminal:
                raise ValueError(f"WACC ({wacc:.2%}) debe ser mayor que crecimiento terminal ({g_terminal:.2%})")
            return ultimo_fcf * (1 + g_terminal) / (wacc - g_terminal)
        raise ValueError(f"Método no reconocido: {metodo}")

    def _descontar_flujos(self, fcfs: List[float], wacc: float) -> List[float]:
        """Descuenta los flujos de caja a valor presente."""
        if len(fcfs) >= 7:
            fcf_proy = fcfs[2:]
        else:
            fcf_proy = fcfs
        descontados = []
        for t, fcf in enumerate(fcf_proy, 1):
            descontados.append(fcf / ((1 + wacc) ** t))
        return descontados

    def _calcular_multiples(self, ev: float, df_er: pd.DataFrame, equity: float, params: dict) -> dict:
        """Calcula los múltiplos de valuación."""
        ingresos_y0 = df_er.loc["Ingresos", "Año 0"]
        ebit_y0 = df_er.loc["EBIT", "Año 0"]
        dya_y0 = df_er.loc["D&A", "Año 0"]
        ebitda_y0 = ebit_y0 + dya_y0
        return {
            "EV / Ventas Año 0": ev / ingresos_y0 if ingresos_y0 else 0,
            "EV / EBITDA Año 0": ev / ebitda_y0 if ebitda_y0 else 0,
            "Valor / Capital Contable": equity / params.get("capital_contable", 1),
        }


class ExcelGenerator:
    """Genera la plantilla Excel de valuación DCF."""

    def __init__(self):
        self.header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        self.header_fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type='solid')
        self.sub_fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type='solid')
        self.sub_font = Font(name='Calibri', bold=True, size=11, color=AZUL_OSCURO)
        self.input_fill = PatternFill(start_color=AMARILLO_INPUT, end_color=AMARILLO_INPUT, fill_type='solid')
        self.output_fill = PatternFill(start_color=VERDE_OUTPUT, end_color=VERDE_OUTPUT, fill_type='solid')
        self.label_font = Font(name='Calibri', size=11, color=COLOR_TEXTO)
        self.val_font = Font(name='Calibri', size=11, bold=True, color=AZUL_OSCURO)
        self.title_font = Font(name='Calibri', bold=True, size=16, color=AZUL_OSCURO)
        self.subtitle_font = Font(name='Calibri', bold=True, size=14, color=AZUL_MEDIO)
        self.border = Border(left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'), top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'))
        self.center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _st(self, cell, font=None, fill=None, alignment=None, border=None, nf=None):
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if nf: cell.number_format = nf

    def _hdr(self, ws, row, vals, sc=1):
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=sc+i, value=v)
            self._st(c, font=self.header_font, fill=self.header_fill, alignment=self.center, border=self.border)

    def _lbl(self, ws, r, c, v):
        cl = ws.cell(row=r, column=c, value=v)
        self._st(cl, font=self.label_font, alignment=self.left, border=self.border)
        return cl

    def _inp(self, ws, r, c, v, nf=None):
        cl = ws.cell(row=r, column=c, value=v)
        self._st(cl, font=self.val_font, fill=self.input_fill, alignment=self.center, border=self.border, nf=nf)
        return cl

    def _frm(self, ws, r, c, f, nf=None):
        cl = ws.cell(row=r, column=c, value=f)
        self._st(cl, font=self.val_font, alignment=self.center, border=self.border, nf=nf)
        return cl

    def _out(self, ws, r, c, v, nf=None):
        cl = ws.cell(row=r, column=c, value=v)
        self._st(cl, font=self.val_font, fill=self.output_fill, alignment=self.center, border=self.border, nf=nf)
        return cl

    def _cols(self, ws, widths):
        for k, v in widths.items():
            ws.column_dimensions[k].width = v

    def generate(self, output_path="assets/plantilla_valuacion.xlsx") -> str:
        """Genera la plantilla Excel completa con 5 hojas interconectadas."""
        wb = Workbook()
        ws1 = wb.active
        self._crear_supuestos(ws1)
        # CAPITAL_TRABAJO se crea ANTES que FLUJO_EFECTIVO para evitar referencias
        # a hojas que aún no existen (openpyxl lo maneja, pero es mejor práctica)
        self._crear_ct(wb)
        self._crear_er(wb)
        self._crear_fe(wb)
        self._crear_val(wb)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        wb.save(output_path)
        return output_path

    def _crear_supuestos(self, ws):
        """Crea la hoja de SUPUESTOS con datos de entrada (sin parámetro wb no usado)."""
        ws.title = "SUPUESTOS"
        ws.sheet_properties.tabColor = AZUL_OSCURO
        self._cols(ws, {'A': 5, 'B': 38, 'C': 20, 'D': 20, 'E': 20, 'F': 20, 'G': 20, 'H': 20, 'I': 20})
        ws.merge_cells('A1:I1')
        self._st(ws.cell(row=1, column=1, value="PLANTILLA DE VALUACIÓN DCF - SUPUESTOS GENERALES"), font=self.title_font, alignment=self.center)
        r = 3
        self._hdr(ws, r, ["", "Concepto", "Valor", "", "", "", "", "", ""])
        r = 4; self._lbl(ws, r, 2, "Empresa"); self._inp(ws, r, 3, SAMPLE_DATA["empresa"])
        r = 5; self._lbl(ws, r, 2, "Fecha de Valuación"); self._inp(ws, r, 3, SAMPLE_DATA["fecha"], nf='DD/MM/YYYY')
        r = 6; self._lbl(ws, r, 2, "Moneda"); self._inp(ws, r, 3, SAMPLE_DATA["moneda"])
        r = 8
        ws.merge_cells(f'A{r}:I{r}')
        self._st(ws.cell(row=r, column=1, value="CRECIMIENTO DE INGRESOS"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 10): ws.cell(row=r, column=c).fill = self.sub_fill
        r = 9; self._hdr(ws, r, ["", "Concepto"] + list(SAMPLE_DATA["crecimiento"].keys()) + [""])
        r = 10; self._lbl(ws, r, 2, "Tasa de Crecimiento")
        for i, a in enumerate(SAMPLE_DATA["crecimiento"].keys()):
            self._inp(ws, r, 3+i, SAMPLE_DATA["crecimiento"][a], nf='0.00%')
        r = 12
        ws.merge_cells(f'A{r}:I{r}')
        self._st(ws.cell(row=r, column=1, value="ESTRUCTURA DE COSTOS (% de Ingresos)"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 10): ws.cell(row=r, column=c).fill = self.sub_fill
        r = 13; self._hdr(ws, r, ["", "Concepto"] + ANOS + [""])
        r = 14; self._lbl(ws, r, 2, "Costo Directo")
        for i, a in enumerate(ANOS): self._inp(ws, r, 3+i, SAMPLE_DATA["costo_directo_pct"].get(a, 0.63), nf='0.00%')
        r = 15; self._lbl(ws, r, 2, "Gastos de Venta")
        for i, a in enumerate(ANOS): self._inp(ws, r, 3+i, SAMPLE_DATA["gastos_venta_pct"].get(a, 0.242), nf='0.00%')
        r = 16; self._lbl(ws, r, 2, "Gastos Administrativos")
        for i, a in enumerate(ANOS): self._inp(ws, r, 3+i, SAMPLE_DATA["gastos_admin_pct"].get(a, 0.019), nf='0.00%')
        r = 17; self._lbl(ws, r, 2, "Depreciación y Amortización (D&A)")
        for i, a in enumerate(ANOS): self._inp(ws, r, 3+i, SAMPLE_DATA["dya_pct"].get(a, 0.025), nf='0.00%')
        r = 19
        ws.merge_cells(f'A{r}:I{r}')
        self._st(ws.cell(row=r, column=1, value="DÍAS DE CAPITAL DE TRABAJO"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 10): ws.cell(row=r, column=c).fill = self.sub_fill
        r = 20; self._hdr(ws, r, ["", "Concepto", "Valor", "", "", "", "", "", ""])
        r = 21; self._lbl(ws, r, 2, "Días Cuentas por Cobrar (CxC)"); self._inp(ws, r, 3, SAMPLE_DATA["dias_cxc"])
        r = 22; self._lbl(ws, r, 2, "Días Inventario"); self._inp(ws, r, 3, SAMPLE_DATA["dias_inventario"])
        r = 23; self._lbl(ws, r, 2, "Días Cuentas por Pagar (CxP)"); self._inp(ws, r, 3, SAMPLE_DATA["dias_cxp"])
        r = 25
        ws.merge_cells(f'A{r}:I{r}')
        self._st(ws.cell(row=r, column=1, value="TASAS E IMPUESTOS"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 10): ws.cell(row=r, column=c).fill = self.sub_fill
        r = 26; self._hdr(ws, r, ["", "Concepto", "Valor", "", "", "", "", "", ""])
        r = 27; self._lbl(ws, r, 2, "ISR + PTU"); self._inp(ws, r, 3, SAMPLE_DATA["isr_ptu"], nf='0.00%')
        r = 28; self._lbl(ws, r, 2, "CapEx (% de Ingresos)"); self._inp(ws, r, 3, SAMPLE_DATA["capex_pct"], nf='0.00%')
        r = 31
        ws.merge_cells(f'A{r}:I{r}')
        self._st(ws.cell(row=r, column=1, value="WACC Y VALOR TERMINAL"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 10): ws.cell(row=r, column=c).fill = self.sub_fill
        r = 32; self._hdr(ws, r, ["", "Concepto", "Valor", "", "", "", "", "", ""])
        r = 33; self._lbl(ws, r, 2, "WACC (Costo Promedio Ponderado de Capital)"); self._inp(ws, r, 3, SAMPLE_DATA["wacc"], nf='0.00%')
        r = 34; self._lbl(ws, r, 2, "Crecimiento Terminal (Perpetuidad)"); self._inp(ws, r, 3, SAMPLE_DATA["crecimiento_terminal"], nf='0.00%')
        r = 35; self._lbl(ws, r, 2, "Método de Valor Terminal"); self._inp(ws, r, 3, SAMPLE_DATA["metodo_tv"])
        r = 37
        ws.merge_cells(f'A{r}:I{r}')
        self._st(ws.cell(row=r, column=1, value="DATOS DE BALANCE GENERAL"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 10): ws.cell(row=r, column=c).fill = self.sub_fill
        r = 38; self._hdr(ws, r, ["", "Concepto", "Valor", "", "", "", "", "", ""])
        r = 39; self._lbl(ws, r, 2, "Deuda Total"); self._inp(ws, r, 3, SAMPLE_DATA["deuda"], nf='#,##0')
        r = 40; self._lbl(ws, r, 2, "Efectivo y Equivalentes"); self._inp(ws, r, 3, SAMPLE_DATA["efectivo"], nf='#,##0')
        r = 41; self._lbl(ws, r, 2, "Capital Contable"); self._inp(ws, r, 3, SAMPLE_DATA["capital_contable"], nf='#,##0')
        ws.freeze_panes = 'A3'

    def _crear_er(self, wb):
        ws = wb.create_sheet("ESTADO_RESULTADOS")
        ws.sheet_properties.tabColor = AZUL_MEDIO
        self._cols(ws, {'A': 5, 'B': 35, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18, 'I': 18})
        ws.merge_cells('A1:I1')
        self._st(ws.cell(row=1, column=1, value="ESTADO DE RESULTADOS PROYECTADO (Miles de MXN)"), font=self.title_font, alignment=self.center)
        self._hdr(ws, 2, ["", "Concepto"] + ANOS + [""])
        
        # Fila 3: Ingresos
        r = 3
        self._lbl(ws, r, 2, "Ingresos")
        self._inp(ws, r, 3, SAMPLE_DATA["ingresos"]["Año -1"], nf='#,##0')
        self._inp(ws, r, 4, SAMPLE_DATA["ingresos"]["Año 0"], nf='#,##0')
        for i, a in enumerate(ANOS_PROY, 5):
            col_ant = get_column_letter(i-1)
            growth_col = get_column_letter(3 + ANOS.index(a))
            self._frm(ws, r, i, f"={col_ant}{r}*(1+'SUPUESTOS'!${growth_col}$10)", nf='#,##0')
        
        # Fila 4: Crecimiento %
        r = 4
        self._lbl(ws, r, 2, "Crecimiento %")
        self._inp(ws, r, 3, SAMPLE_DATA["crecimiento"]["Año -1"], nf='0.00%')
        # Año 0: fórmula (dinámica)
        self._frm(ws, r, 4, "='SUPUESTOS'!$D$10", nf='0.00%')
        for i, a in enumerate(ANOS_PROY, 5):
            growth_col = get_column_letter(3 + ANOS.index(a))
            self._frm(ws, r, i, f"='SUPUESTOS'!${growth_col}$10", nf='0.00%')
        
        # Fila 6: Costo Directo (corregido: $C$14)
        r = 6
        self._lbl(ws, r, 2, "Costo Directo")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            if a in SAMPLE_DATA["ingresos"]:
                ing = SAMPLE_DATA["ingresos"].get(a, 0)
                pct = SAMPLE_DATA["costo_directo_pct"].get(a, 0.636)
                self._inp(ws, r, i, ing * pct, nf='#,##0')
            else:
                self._frm(ws, r, i, f"={cl}3*'SUPUESTOS'!$C$14", nf='#,##0')
        
        # Fila 7: Gastos de Venta ($C$15)
        r = 7
        self._lbl(ws, r, 2, "Gastos de Venta")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            if a in SAMPLE_DATA["ingresos"]:
                ing = SAMPLE_DATA["ingresos"].get(a, 0)
                pct = SAMPLE_DATA["gastos_venta_pct"].get(a, 0.242)
                self._inp(ws, r, i, ing * pct, nf='#,##0')
            else:
                self._frm(ws, r, i, f"={cl}3*'SUPUESTOS'!$C$15", nf='#,##0')
        
        # Fila 8: Gastos Administrativos ($C$16)
        r = 8
        self._lbl(ws, r, 2, "Gastos Administrativos")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            if a in SAMPLE_DATA["ingresos"]:
                ing = SAMPLE_DATA["ingresos"].get(a, 0)
                pct = SAMPLE_DATA["gastos_admin_pct"].get(a, 0.019)
                self._inp(ws, r, i, ing * pct, nf='#,##0')
            else:
                self._frm(ws, r, i, f"={cl}3*'SUPUESTOS'!$C$16", nf='#,##0')
        
        # Fila 9: D&A ($C$17)
        r = 9
        self._lbl(ws, r, 2, "Depreciación y Amortización")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            if a in SAMPLE_DATA["ingresos"]:
                ing = SAMPLE_DATA["ingresos"].get(a, 0)
                pct = SAMPLE_DATA["dya_pct"].get(a, 0.025)
                self._inp(ws, r, i, ing * pct, nf='#,##0')
            else:
                self._frm(ws, r, i, f"={cl}3*'SUPUESTOS'!$C$17", nf='#,##0')
        
        # Fila 11: Total Costos y Gastos (sin cambios)
        r = 11
        self._lbl(ws, r, 2, "Total Costos y Gastos")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"=SUM({cl}6:{cl}9)", nf='#,##0')
        
        # Fila 13: EBIT (sin cambios)
        r = 13
        self._lbl(ws, r, 2, "Utilidad de Operación (EBIT)")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}3-{cl}11", nf='#,##0')
        
        # Fila 14: Margen EBIT % (sin cambios)
        r = 14
        self._lbl(ws, r, 2, "Margen EBIT %")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}13/{cl}3", nf='0.00%')
        
        # Fila 16: ISR + PTU ($C$27)
        r = 16
        self._lbl(ws, r, 2, "ISR + PTU")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}13*'SUPUESTOS'!$C$27", nf='#,##0')
        
        # Fila 18: NOPAT (sin cambios)
        r = 18
        self._lbl(ws, r, 2, "NOPAT (Utilidad Neta Operativa)")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}13-{cl}16", nf='#,##0')
        
        # Fila 19: Margen Neto % (sin cambios)
        r = 19
        self._lbl(ws, r, 2, "Margen Neto %")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}18/{cl}3", nf='0.00%')
        
        # Fila 21: Utilidad Neta (sin cambios)
        r = 21
        self._lbl(ws, r, 2, "Utilidad Neta")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}18", nf='#,##0')
        
        ws.freeze_panes = 'A3'

    def _crear_fe(self, wb):
        ws = wb.create_sheet("FLUJO_EFECTIVO")
        ws.sheet_properties.tabColor = "27AE60"
        self._cols(ws, {'A': 5, 'B': 40, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18, 'I': 18})
        ws.merge_cells('A1:I1')
        self._st(ws.cell(row=1, column=1, value="FLUJO DE CAJA LIBRE PROYECTADO (Miles de MXN)"), font=self.title_font, alignment=self.center)
        self._hdr(ws, 2, ["", "Concepto"] + ANOS + [""])
        
        # Fila 3: NOPAT
        r = 3
        self._lbl(ws, r, 2, "NOPAT")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='ESTADO_RESULTADOS'!{cl}18", nf='#,##0')
        
        # Fila 5: D&A
        r = 5
        self._lbl(ws, r, 2, "(+) Depreciación y Amortización")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='ESTADO_RESULTADOS'!{cl}9", nf='#,##0')
        
        # Fila 7: Δ Capital de Trabajo
        r = 7
        self._lbl(ws, r, 2, "(-) Δ Capital de Trabajo")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='CAPITAL_TRABAJO'!{cl}12", nf='#,##0')
        
        # Fila 9: CapEx (corregido: $C$28)
        r = 9
        self._lbl(ws, r, 2, "(-) CapEx")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='ESTADO_RESULTADOS'!{cl}3*'SUPUESTOS'!$C$28", nf='#,##0')
        
        # Fila 11: FCF
        r = 11
        self._lbl(ws, r, 2, "Flujo de Caja Libre (FCF)")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}3+{cl}5-{cl}7-{cl}9", nf='#,##0')
        
        # Fila 13: Factor de Descuento (corregido: $C$33)
        r = 13
        self._lbl(ws, r, 2, "Factor de Descuento (WACC)")
        for i, a in enumerate(ANOS, 3):
            t = i - 2
            if t <= 2:
                self._frm(ws, r, i, 1.0, nf='0.0000')
            else:
                t_real = t - 2
                self._frm(ws, r, i, f"=1/(1+'SUPUESTOS'!$C$33)^{t_real}", nf='0.0000')
        
        # Fila 15: FCF Descontado
        r = 15
        self._lbl(ws, r, 2, "FCF Descontado")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}11*{cl}13", nf='#,##0')
        
        # Fila 17: Valor Terminal (corregido: $C$33, $C$34, $C$35)
        r = 17
        self._lbl(ws, r, 2, "Valor Terminal (Perpetuidad)")
        ws.merge_cells(f'C{r}:H{r}')
        self._frm(ws, r, 3,
            f'=IF(\'SUPUESTOS\'!$C$35="Growth",I11*(1+\'SUPUESTOS\'!$C$34)/(\'SUPUESTOS\'!$C$33-\'SUPUESTOS\'!$C$34),'
            f'IF(\'SUPUESTOS\'!$C$35="Zero",I11/\'SUPUESTOS\'!$C$33,0))', nf='#,##0')
        
        # Fila 19: TV Descontado
        r = 19
        self._lbl(ws, r, 2, "Valor Terminal Descontado")
        ws.merge_cells(f'C{r}:H{r}')
        self._frm(ws, r, 3, "=C17*I13", nf='#,##0')
        
        ws.freeze_panes = 'A3'

    def _crear_val(self, wb):
        ws = wb.create_sheet("VALUACION")
        ws.sheet_properties.tabColor = NARANJA_ACENTO
        self._cols(ws, {'A': 5, 'B': 40, 'C': 22, 'D': 22})
        ws.merge_cells('A1:D1')
        self._st(ws.cell(row=1, column=1, value="RESUMEN DE VALUACIÓN (Miles de MXN)"), font=self.title_font, alignment=self.center)
        self._hdr(ws, 3, ["", "Concepto", "Valor", ""])
        
        # Enterprise Value
        r = 5
        self._lbl(ws, r, 2, "Enterprise Value (EV)")
        self._out(ws, r, 3, f"=SUM('FLUJO_EFECTIVO'!E15:I15)+'FLUJO_EFECTIVO'!C19", nf='#,##0')
        
        # Deuda ($C$39)
        r = 7
        self._lbl(ws, r, 2, "(-) Deuda Total")
        self._frm(ws, r, 3, "='SUPUESTOS'!$C$39", nf='#,##0')
        
        # Efectivo ($C$40)
        r = 9
        self._lbl(ws, r, 2, "(+) Efectivo y Equivalentes")
        self._frm(ws, r, 3, "='SUPUESTOS'!$C$40", nf='#,##0')
        
        # Equity Value
        r = 11
        self._lbl(ws, r, 2, "Valor del Equity")
        self._out(ws, r, 3, "=C5-C7+C9", nf='#,##0')
        
        # Múltiplos
        r = 14
        ws.merge_cells(f'A{r}:D{r}')
        self._st(ws.cell(row=r, column=1, value="MÚLTIPLOS DE VALUACIÓN"), font=self.subtitle_font, fill=self.sub_fill, alignment=self.left)
        for c in range(2, 5):
            ws.cell(row=r, column=c).fill = self.sub_fill
        
        r = 15
        self._hdr(ws, r, ["", "Indicador", "Valor", ""])
        
        r = 17
        self._lbl(ws, r, 2, "EV / Ventas Año 0")
        self._out(ws, r, 3, "=C5/'ESTADO_RESULTADOS'!D3", nf='0.00"x"')
        
        r = 19
        self._lbl(ws, r, 2, "EV / EBITDA Año 0")
        self._out(ws, r, 3, "=C5/('ESTADO_RESULTADOS'!D13+'ESTADO_RESULTADOS'!D9)", nf='0.00"x"')
        
        r = 21
        self._lbl(ws, r, 2, "Valor / Capital Contable")
        self._out(ws, r, 3, "=C11/'SUPUESTOS'!$C$41", nf='0.00"x"')
        
        ws.freeze_panes = 'A3'

    def _crear_ct(self, wb):
        ws = wb.create_sheet("CAPITAL_TRABAJO")
        ws.sheet_properties.tabColor = "8E44AD"
        self._cols(ws, {'A': 5, 'B': 40, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18, 'I': 18})
        ws.merge_cells('A1:I1')
        self._st(ws.cell(row=1, column=1, value="DETALLE DE CAPITAL DE TRABAJO (Miles de MXN)"), font=self.title_font, alignment=self.center)
        self._hdr(ws, 2, ["", "Concepto"] + ANOS + [""])
        
        # Cuentas por Cobrar ($C$21)
        r = 4
        self._lbl(ws, r, 2, "Cuentas por Cobrar (CxC)")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='SUPUESTOS'!$C$21/365*'ESTADO_RESULTADOS'!{cl}3", nf='#,##0')
        
        # Inventario ($C$22)
        r = 6
        self._lbl(ws, r, 2, "Inventario")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='SUPUESTOS'!$C$22/365*'ESTADO_RESULTADOS'!{cl}6", nf='#,##0')
        
        # Cuentas por Pagar ($C$23)
        r = 8
        self._lbl(ws, r, 2, "Cuentas por Pagar (CxP)")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"='SUPUESTOS'!$C$23/365*('ESTADO_RESULTADOS'!{cl}6+'ESTADO_RESULTADOS'!{cl}7+'ESTADO_RESULTADOS'!{cl}8)", nf='#,##0')
        
        # Working Capital
        r = 10
        self._lbl(ws, r, 2, "Working Capital (CxC + Inv - CxP)")
        for i, a in enumerate(ANOS, 3):
            cl = get_column_letter(i)
            self._frm(ws, r, i, f"={cl}4+{cl}6-{cl}8", nf='#,##0')
        
        # Δ Working Capital
        r = 12
        self._lbl(ws, r, 2, "Δ Working Capital")
        for i, a in enumerate(ANOS, 3):
            if i == 3:
                self._frm(ws, r, i, 0, nf='#,##0')
            else:
                ca = get_column_letter(i)
                cb = get_column_letter(i-1)
                self._frm(ws, r, i, f"={ca}10-{cb}10", nf='#,##0')
        
        ws.freeze_panes = 'A3'


def generar_parametros(ingresos_base=741120,
                        crec_y1=0.05, crec_y2=0.05, crec_y3=0.05, crec_y4=0.05, crec_y5=0.05,
                        costo_directo_pct=0.636,
                        gastos_venta_pct=0.242, gastos_admin_pct=0.019, dya_pct=0.025,
                        dias_cxc=45, dias_inventario=20, dias_cxp=69,
                        isr_ptu=0.008, capex_pct=0.025, wacc=0.175,
                        crecimiento_terminal=0.0, metodo_tv="Growth",
                        deuda=54739, efectivo=7411, capital_contable=218956):
    """Genera el diccionario de parámetros para el motor DCF."""
    return {
        "ingresos_base": ingresos_base,
        "ingresos_historicos": {"Año -1": 622590},
        "crecimiento": {"Año -1": 0.163, "Año 0": 0.190, "Año 1": crec_y1, "Año 2": crec_y2, "Año 3": crec_y3, "Año 4": crec_y4, "Año 5": crec_y5},
        "costo_directo_pct": {a: costo_directo_pct for a in ANOS},
        "gastos_venta_pct": {a: gastos_venta_pct for a in ANOS},
        "gastos_admin_pct": {a: gastos_admin_pct for a in ANOS},
        "dya_pct": {a: dya_pct for a in ANOS},
        "dias_cxc": dias_cxc, "dias_inventario": dias_inventario, "dias_cxp": dias_cxp,
        "isr_ptu": isr_ptu, "capex_pct": capex_pct,
        "wacc": wacc, "crecimiento_terminal": crecimiento_terminal, "metodo_tv": metodo_tv,
        "deuda": deuda, "efectivo": efectivo, "capital_contable": capital_contable,
    }


def plot_fcf(fcfs, terminal_value):
    """Gráfica de barras: Flujos de Caja Libre proyectados + TV."""
    labels = ["Año 1", "Año 2", "Año 3", "Año 4", "Año 5", "Valor Terminal"]
    valores = list(fcfs[-5:]) + [terminal_value]
    fig = go.Figure(data=[go.Bar(x=labels, y=valores,
        text=[f"${v:,.0f}" for v in valores], textposition='outside',
        marker_color=['#2E86C1']*5 + ['#F39C12'])])
    fig.update_layout(title="Flujos de Caja Libre Proyectados", yaxis_title="Miles de MXN",
        template="plotly_white", height=400)
    return fig


def plot_composicion(fcf_pv_sum, tv_pv):
    """Gráfica de dona: composición del valor (FCF vs TV)."""
    fig = go.Figure(data=[go.Pie(labels=['Flujos Años 1-5', 'Valor Terminal'],
        values=[fcf_pv_sum, tv_pv], hole=0.4,
        marker_colors=['#2E86C1', '#F39C12'],
        textinfo='label+percent', texttemplate='%{label}<br>%{percent} ($%{value:,.0f})')])
    fig.update_layout(title="Composición del Enterprise Value", height=400)
    return fig


def plot_sensibilidad(engine, params_base):
    """Heatmap de sensibilidad: WACC vs Crecimiento Terminal."""
    wacc_range = np.arange(0.08, 0.28, 0.02)
    g_range = np.arange(0.0, 0.06, 0.01)
    matrix = np.zeros((len(wacc_range), len(g_range)))
    for i, w in enumerate(wacc_range):
        for j, g in enumerate(g_range):
            p = params_base.copy()
            p["wacc"] = w
            p["crecimiento_terminal"] = g
            # FIX ERROR 7: Capturar solo ValueError, no todas las excepciones
            try:
                matrix[i][j] = engine.calcular_valuacion(p)["enterprise_value"]
            except ValueError:
                matrix[i][j] = 0
    fig = go.Figure(data=go.Heatmap(z=matrix, x=[f"{g:.1%}" for g in g_range],
        y=[f"{w:.1%}" for w in wacc_range],
        text=np.round(matrix, 0), texttemplate="$%{text:,.0f}",
        colorscale='RdYlGn', hovertemplate='WACC: %{y}<br>Crec. Terminal: %{x}<br>EV: $%{z:,.0f}<extra></extra>'))
    fig.update_layout(title="Sensibilidad: WACC vs Crecimiento Terminal",
        xaxis_title="Crecimiento Terminal", yaxis_title="WACC", height=450)
    return fig


def apply_custom_css():
    """Aplica estilos CSS personalizados a la aplicación."""
    st.markdown("""
    <style>
        .stApp { background-color: #F8F9FA; }
        h1, h2, h3 { color: #1B2A4A !important; font-family: 'Calibri', sans-serif; }
        .stButton button { background-color: #2E86C1; color: white; border: none; border-radius: 6px; padding: 0.5rem 1rem; font-weight: 600; transition: all 0.3s ease; }
        .stButton button:hover { background-color: #1B2A4A; transform: translateY(-2px); box-shadow: 0 4px 8px rgba(0,0,0,0.1); }
        .stMetric { background-color: white; padding: 1.5rem; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #2E86C1; }
        .stMetric label { color: #7F8C8D; font-size: 0.9rem; }
        .stMetric .css-1wivap2 { color: #1B2A4A; font-size: 1.8rem; font-weight: 700; }
        .stDataFrame { background-color: white; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
        .stTabs [data-baseweb="tab-list"] { gap: 8px; }
        .stTabs [data-baseweb="tab"] { border-radius: 8px 8px 0 0; padding: 10px 24px; font-weight: 600; }
        .stTabs [aria-selected="true"] { background-color: #1B2A4A !important; color: white !important; }
        .stAlert { border-radius: 8px; border-left: 4px solid #2E86C1; }
        .stFileUploader { background-color: white; padding: 1rem; border-radius: 8px; border: 2px dashed #BDC3C7; }
        th { background-color: #1B2A4A !important; color: white !important; padding: 12px !important; }
        td { padding: 10px !important; border-bottom: 1px solid #ECF0F1; }
    </style>
    """, unsafe_allow_html=True)


def init_session():
    """Inicializa las variables de estado de sesión de Streamlit."""
    if "params" not in st.session_state:
        st.session_state.params = None
    if "resultado" not in st.session_state:
        st.session_state.resultado = None
    if "datos_cargados" not in st.session_state:
        st.session_state.datos_cargados = False


def tab_inicio():
    """Renderiza la pestaña de Inicio/Bienvenida."""
    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown("""
        ## 🏢 Valuador de Empresas por Flujos Descontados (DCF)
        **Metodología profesional de valuación** utilizada por los principales bancos de inversión y firmas de consultoría.
        ---
        ### 📋 ¿Cómo funciona?
        1. **Descarga la plantilla Excel** con datos de prueba precargados
        2. **Llena los datos** de la empresa que quieres valuar
        3. **Carga el archivo** en la aplicación
        4. **Ajusta parámetros** en el simulador interactivo
        5. **Visualiza resultados** con gráficas profesionales
        """)
    with col2:
        st.markdown("### Comenzar")
        if st.button("📥 Descargar Plantilla Excel", use_container_width=True):
            with st.spinner("Generando plantilla..."):
                gen = ExcelGenerator()
                ruta = gen.generate("assets/plantilla_valuacion.xlsx")
                with open(ruta, "rb") as f:
                    data = f.read()
                st.download_button(label="✅ Guardar Plantilla", data=data,
                    file_name="plantilla_valuacion_dcf.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        st.markdown("---")
        archivo = st.file_uploader("📤 Cargar plantilla llena", type=["xlsx"],
            help="Sube el archivo Excel que descargaste y llenaste")
        if archivo:
            with st.spinner("Procesando archivo..."):
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(archivo.getvalue())
                    ruta_tmp = tmp.name
                try:
                    wb = load_workbook(ruta_tmp, data_only=True)
                    ws = wb["SUPUESTOS"]
                    st.session_state.params = generar_parametros(
                        ingresos_base=safe_float(ws["C3"].value, 741120),  # Año 0 ingresos está en C3 (en SUPUESTOS)
                        crec_y1=safe_float(ws["E10"].value, 0.05),        # Crecimiento Año 1 está en E10
                        dias_cxc=safe_float(ws["C21"].value, 45),         # Días CxC en C21
                        dias_inventario=safe_float(ws["C22"].value, 20),  # Días Inventario en C22
                        dias_cxp=safe_float(ws["C23"].value, 69),         # Días CxP en C23
                        wacc=safe_float(ws["C33"].value, 0.175),          # WACC en C33
                    )
                    st.session_state.datos_cargados = True
                    st.success("✅ Datos cargados exitosamente. Ve al Simulador.")
                    os.unlink(ruta_tmp)
                except Exception as e:
                    st.error(f"❌ Error al leer el archivo: {e}")
                    if os.path.exists(ruta_tmp):
                        os.unlink(ruta_tmp)
        st.markdown("---")
        if st.button("🔄 Cargar Datos de Prueba", use_container_width=True):
            st.session_state.params = generar_parametros()
            st.session_state.datos_cargados = True
            st.success("✅ Datos de prueba cargados. Ve al Simulador.")


def tab_simulador():
    """Renderiza la pestaña del Simulador interactivo."""
    if not st.session_state.datos_cargados:
        st.warning("⚠️ Primero carga los datos en la pestaña de INICIO")
        return
    st.markdown("# 💰 Simulador de Valuación")
    kpi_cols = st.columns(3)
    params_actuales = st.session_state.params
    
    # Función para obtener valor numérico seguro
    def get_float(key, default):
        val = params_actuales.get(key, default)
        try:
            return float(val)
        except (ValueError, TypeError):
            return float(default)

    with st.sidebar:
        st.markdown("## ⚙️ Parámetros")
        wacc = st.slider("WACC", 0.05, 0.30, get_float("wacc", 0.175), 0.005, format="%.1f%%")
        st.markdown("---")
        st.markdown("### 📈 Crecimiento de Ingresos")
        g1 = st.slider("Crecimiento Año 1", -0.10, 0.50, get_float("crecimiento_Año1", 0.05), 0.005, format="%.1f%%")
        g2 = st.slider("Crecimiento Año 2", -0.10, 0.50, get_float("crecimiento_Año2", 0.05), 0.005, format="%.1f%%")
        g3 = st.slider("Crecimiento Año 3", -0.10, 0.50, get_float("crecimiento_Año3", 0.05), 0.005, format="%.1f%%")
        g4 = st.slider("Crecimiento Año 4", -0.10, 0.50, get_float("crecimiento_Año4", 0.05), 0.005, format="%.1f%%")
        g5 = st.slider("Crecimiento Año 5", -0.10, 0.50, get_float("crecimiento_Año5", 0.05), 0.005, format="%.1f%%")
        st.markdown("---")
        st.markdown("### 🔄 Capital de Trabajo")
        dcxc = st.slider("Días CxC", 0, 120, int(get_float("dias_cxc", 45)), step=1)
        dinv = st.slider("Días Inventario", 0, 90, int(get_float("dias_inventario", 20)), step=1)
        dcxp = st.slider("Días CxP", 0, 120, int(get_float("dias_cxp", 69)), step=1)
        st.markdown("---")
        st.markdown("### 💰 Costos")
        cd = st.slider("Costo Directo %", 0.30, 0.80, get_float("costo_directo_pct", 0.636), 0.005, format="%.1f%%")
        isr = st.slider("ISR + PTU %", 0.0, 0.35, get_float("isr_ptu", 0.008), 0.005, format="%.1f%%")
        st.markdown("---")
        st.markdown("### 🏢 Valor Terminal")
        gt = st.slider("Crecimiento Terminal", 0.0, 0.10, get_float("crecimiento_terminal", 0.0), 0.005, format="%.1f%%")
        met = st.selectbox("Método TV", ["Growth", "Zero", "None"], index=0)
    params = generar_parametros(
        ingresos_base=st.session_state.params.get("ingresos_base", 741120),
        crec_y1=g1, crec_y2=g2, crec_y3=g3, crec_y4=g4, crec_y5=g5,
        costo_directo_pct=cd, isr_ptu=isr,
        dias_cxc=dcxc, dias_inventario=dinv, dias_cxp=dcxp,
        wacc=wacc, crecimiento_terminal=gt, metodo_tv=met,
    )
    engine = DCFEngine()
    resultado = engine.calcular_valuacion(params)
    st.session_state.resultado = resultado
    with kpi_cols[0]:
        st.metric("💰 Valor del Equity", f"${resultado['equity_value']:,.0f}")
    with kpi_cols[1]:
        st.metric("🏢 Enterprise Value", f"${resultado['enterprise_value']:,.0f}")
    with kpi_cols[2]:
        st.metric("📊 EV / EBITDA", f"{resultado['multiples'].get('EV / EBITDA Año 0', 0):.2f}x")
    st.markdown("---")
    colg1, colg2 = st.columns(2)
    with colg1:
        st.plotly_chart(plot_fcf(resultado["fcfs"], resultado["terminal_value"]), use_container_width=True)
    with colg2:
        fcf_pv = sum(resultado["fcf_descontados"])
        tv_pv = resultado["tv_descontado"]
        st.plotly_chart(plot_composicion(fcf_pv, tv_pv), use_container_width=True)
    st.markdown("---")
    colg3, _ = st.columns([2, 1])
    with colg3:
        st.plotly_chart(plot_sensibilidad(engine, params), use_container_width=True)
    with st.expander("📋 Ver proyección detallada"):
        st.dataframe(resultado["flujo_efectivo"], use_container_width=True)


def tab_estados():
    """Renderiza la pestaña de Estados Financieros."""
    if not st.session_state.datos_cargados:
        st.warning("⚠️ Primero carga los datos en la pestaña de INICIO")
        return
    st.markdown("# 📊 Estados Financieros Proyectados")
    if st.session_state.resultado:
        r = st.session_state.resultado
        # FIX ERROR 2: st.tats -> st.tabs
        t1, t2, t3 = st.tabs(["Estado de Resultados", "Flujo de Efectivo", "Capital de Trabajo"])
        with t1:
            st.dataframe(r["estado_resultados"], use_container_width=True)
        with t2:
            st.dataframe(r["flujo_efectivo"], use_container_width=True)
        with t3:
            st.dataframe(r["capital_trabajo"], use_container_width=True)
    else:
        st.info("Ejecuta la valuación en el Simulador primero")


def tab_reportes():
    """Renderiza la pestaña de Reportes y Exportación."""
    if not st.session_state.datos_cargados:
        st.warning("⚠️ Primero carga los datos en la pestaña de INICIO")
        return
    st.markdown("# 📋 Reporte de Valuación")
    if st.session_state.resultado:
        r = st.session_state.resultado
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("### 📄 Resumen Ejecutivo")
            df = pd.DataFrame({
                "Concepto": ["Valor del Equity", "Enterprise Value (EV)", "Deuda Total", "Efectivo", "WACC", "Crec. Terminal", "Método TV"],
                "Valor": [f"${r['equity_value']:,.0f}", f"${r['enterprise_value']:,.0f}", "$54,739", "$7,411", f"{r['wacc']:.1%}", f"{r['crecimiento_terminal']:.1%}", r['metodo_tv']]
            })
            st.table(df)
        with c2:
            st.markdown("### 📊 Múltiplos")
            for k, v in r['multiples'].items():
                st.metric(k, f"{v:.2f}x")
        st.markdown("---")
        st.markdown("### 💾 Exportar Resultados")
        if st.button("📥 Exportar a Excel", use_container_width=True):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                r["estado_resultados"].to_excel(writer, sheet_name="Estado_Resultados")
                r["flujo_efectivo"].to_excel(writer, sheet_name="Flujo_Efectivo")
                r["capital_trabajo"].to_excel(writer, sheet_name="Capital_Trabajo")
                pd.DataFrame({
                    "Métrica": ["Enterprise Value", "Equity Value", "WACC"],
                    "Valor": [r["enterprise_value"], r["equity_value"], r["wacc"]]
                }).to_excel(writer, sheet_name="Resumen", index=False)
            st.download_button("✅ Descargar", data=output.getvalue(),
                file_name="resultado_valuacion.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
    else:
        st.info("Ejecuta la valuación en el Simulador primero")


def main():
    """Función principal de la aplicación. Configura layout, sidebar y pestañas."""
    init_session()
    apply_custom_css()
    st.markdown("""
    <div style="text-align: center; padding: 1rem;">
        <h1 style="color: #1B2A4A;">🏢 Valuador DCF</h1>
        <p style="color: #7F8C8D; font-size: 1.1rem;">Simulador de Valuación por Flujos Descontados</p>
    </div>
    """, unsafe_allow_html=True)
    tabs = st.tabs(["🏠 Inicio", "💰 Simulador", "📊 Estados", "📋 Reportes"])
    with tabs[0]:
        tab_inicio()
    with tabs[1]:
        tab_simulador()
    with tabs[2]:
        tab_estados()
    with tabs[3]:
        tab_reportes()
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #95A5A6; font-size: 0.8rem;">
        <p>Valuador DCF v1.0 | Metodología de Flujos de Caja Descontados (DCF)</p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()